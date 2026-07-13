# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io

def clean_excel_data(file_obj, filename: str):
    """
    数据清洗：去重、去全空行，并返回处理后的内存文件对象及统计信息
    """
    if filename.lower().endswith('.csv'):
        df = pd.read_csv(file_obj)
    else:
        df = pd.read_excel(file_obj)
        
    original_rows = len(df)
    
    # 1. 删除全空行
    df = df.dropna(how='all')
    # 2. 填充空值
    df = df.fillna("N/A")
    # 3. 删除完全重复的行
    df = df.drop_duplicates()
    
    cleaned_rows = len(df)
    
    # 写入内存
    output = io.BytesIO()
    if filename.lower().endswith('.csv'):
        df.to_csv(output, index=False, encoding='utf-8-sig')
    else:
        df.to_excel(output, index=False)
        
    output.seek(0)
    return output, original_rows, cleaned_rows

def format_excel_style(file_obj):
    """
    格式统一：设置字体、居中对齐、添加边框、自适应列宽
    """
    # openpyxl 只能处理 xlsx，如果传入的是已经过 Pandas 处理的 BytesIO，直接加载
    wb = load_workbook(file_obj)
    ws = wb.active
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='微软雅黑', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 4
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
